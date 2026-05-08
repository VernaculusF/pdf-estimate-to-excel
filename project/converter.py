"""
Excel export helpers.
"""

import os
import re
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import DEFAULT_SHEET_NAME, RAW_ESTIMATE_COLUMNS

_ILLEGAL_XML_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufffe\uffff]"
)


_MAX_CELL_CHARS = 32767


def _safe_value(value):
    if value is None:
        return value
    if isinstance(value, float):
        import math
        if math.isnan(value) or math.isinf(value):
            return ""
    if isinstance(value, str):
        cleaned = _ILLEGAL_XML_RE.sub("", value)
        if len(cleaned) > _MAX_CELL_CHARS:
            cleaned = cleaned[:_MAX_CELL_CHARS]
        return cleaned
    return value


class SmetaConverter:
    """Save extracted estimate data to formatted Excel workbooks."""

    def __init__(self):
        self.sheet_name = DEFAULT_SHEET_NAME
        self.header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def save_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: Optional[str] = None,
    ) -> str:
        if df is None:
            raise ValueError("DataFrame is None")

        sheet = sheet_name or self.sheet_name
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=_safe_value(value))
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    cell.alignment = self.left_alignment
                cell.border = self.border

        self._auto_adjust_columns(ws)
        ws.freeze_panes = "A2"
        wb.save(output_path)
        return output_path

    def save_raw_estimate_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: Optional[str] = None,
    ) -> str:
        if df is None:
            raise ValueError("DataFrame is None")

        sheet = sheet_name or self.sheet_name
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        max_cols = min(len(df.columns), 14)
        num_cols = max_cols
        self._write_raw_estimate_headers(ws, num_cols)

        for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
            for col_index, value in enumerate(row[:max_cols], start=1):
                cell = ws.cell(row=row_index, column=col_index, value=_safe_value(value))
                cell.border = self.border
                cell.alignment = self.left_alignment

        self._merge_summary_rows(ws, num_cols)
        self._apply_raw_estimate_layout(ws, num_cols)
        wb.save(output_path)
        return output_path

    def _merge_summary_rows(self, ws, num_cols: int) -> None:
        """Merge cells A:G for rows where text is only in col 1 and cols 2-7 are empty.

        Detection is structure-based: any row with content in column 1 and
        empty columns 2 through 7 is treated as a summary/section row whose
        first 7 cells should be merged — no keyword list needed.
        """
        if num_cols < 8:
            return
        merge_end = min(7, num_cols)
        if merge_end <= 1:
            return
        for row_idx in range(5, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            text = str(cell_a.value or "").strip()
            if not text:
                continue
            cols_2_to_7_empty = all(
                not str(ws.cell(row=row_idx, column=c).value or "").strip()
                for c in range(2, merge_end + 1)
            )
            if not cols_2_to_7_empty:
                continue
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=merge_end,
            )
            cell_a.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True,
            )

    def _write_raw_estimate_headers(self, ws, num_cols: int = 14) -> None:
        row1 = [
            RAW_ESTIMATE_COLUMNS[0],
            RAW_ESTIMATE_COLUMNS[1],
            RAW_ESTIMATE_COLUMNS[2],
            RAW_ESTIMATE_COLUMNS[3],
            "Кол.",
            "",
            "Сметная стоимость в текущих (прогнозных) ценах, руб.",
            "",
            "",
            "",
            "",
            "",
            RAW_ESTIMATE_COLUMNS[12],
            RAW_ESTIMATE_COLUMNS[13],
        ]
        row2 = [
            "",
            "",
            "",
            "",
            "на ед.",
            "всего",
            "на ед.",
            "общая",
            "В том числе",
            "",
            "",
            "",
            "",
            "",
        ]
        row3 = [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            RAW_ESTIMATE_COLUMNS[8],
            RAW_ESTIMATE_COLUMNS[9],
            RAW_ESTIMATE_COLUMNS[10],
            RAW_ESTIMATE_COLUMNS[11],
            "",
            "",
        ]
        row4 = [str(index) for index in range(1, 15)]

        for row_index, row in enumerate((row1, row2, row3, row4), start=1):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.border = self.border
                cell.alignment = self.center_alignment
                if row_index < 4:
                    cell.font = Font(bold=True)

        merge_ranges = [
            "A1:A3",
            "B1:B3",
            "C1:C3",
            "D1:D3",
            "E1:F1",
            "G1:L1",
            "I2:L2",
            "M1:M3",
            "N1:N3",
        ]
        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

    def _apply_raw_estimate_layout(self, ws, num_cols: int = 14) -> None:
        widths = {
            "A": 6,
            "B": 16,
            "C": 52,
            "D": 10,
            "E": 12,
            "F": 12,
            "G": 10,
            "H": 10,
            "I": 10,
            "J": 10,
            "K": 10,
            "L": 10,
            "M": 10,
            "N": 10,
        }
        for column_letter, width in widths.items():
            ws.column_dimensions[column_letter].width = width

        ws.freeze_panes = "A5"
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 24

    def _auto_adjust_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
