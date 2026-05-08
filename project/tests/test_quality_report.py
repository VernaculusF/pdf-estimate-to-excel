import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from converter import SmetaConverter
from quality_report import append_source_text_sheet, build_text_metrics, extract_excel_text


class QualityReportTests(unittest.TestCase):
    def test_build_text_metrics_counts_preserved_visible_characters(self):
        source = "АБВ 123\nЦена: 45,67"
        excel = "абв123 цена:45,67"

        metrics = build_text_metrics(source, excel)

        self.assertEqual(metrics["source_visible_chars"], 16)
        self.assertEqual(metrics["excel_visible_chars"], 16)
        self.assertEqual(metrics["matched_visible_chars"], 16)
        self.assertEqual(metrics["lost_visible_chars"], 0)
        self.assertEqual(metrics["preservation_percent"], 100.0)

    def test_build_text_metrics_reports_lost_characters(self):
        metrics = build_text_metrics("abcdef", "abdf")

        self.assertEqual(metrics["source_visible_chars"], 6)
        self.assertEqual(metrics["matched_visible_chars"], 4)
        self.assertEqual(metrics["lost_visible_chars"], 2)
        self.assertEqual(metrics["preservation_percent"], 66.67)

    def test_converter_preserves_5000_character_cell(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "long_text.xlsx"
            text = ("АБВ123 " * 715)[:5000]

            SmetaConverter().save_to_excel(pd.DataFrame({"Текст": [text]}), str(output_path))

            workbook = load_workbook(output_path)
            saved = workbook.active["A2"].value
            self.assertEqual(len(saved), 5000)
            self.assertEqual(saved, text)

    def test_extract_excel_text_reads_all_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "cells.xlsx"
            df = pd.DataFrame({"A": ["Первая строка"], "B": ["Вторая строка"]})
            SmetaConverter().save_to_excel(df, str(output_path))

            text = extract_excel_text(str(output_path))

            self.assertIn("Первая строка", text)
            self.assertIn("Вторая строка", text)

    def test_append_source_text_sheet_preserves_long_source_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "source_text.xlsx"
            source_text = ("Полный текст PDF 123 " * 250)[:5000]
            SmetaConverter().save_to_excel(pd.DataFrame({"A": ["таблица"]}), str(output_path))

            append_source_text_sheet(str(output_path), source_text, "pdf_text_layer")

            workbook = load_workbook(output_path)
            self.assertIn("Source Text", workbook.sheetnames)
            self.assertEqual(workbook["Source Text"]["A3"].value, source_text)


if __name__ == "__main__":
    unittest.main()
