import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from converter import SmetaConverter


class ConverterLayoutTests(unittest.TestCase):
    def test_save_raw_estimate_to_excel_writes_merged_header(self):
        converter = SmetaConverter()

        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "raw.xlsx"
            df = pd.DataFrame(
                [["2", "CODE", "Description", "шт", "", "20", "496,82", "9936,4", "", "", "", "9936,4", "", ""]],
                columns=[f"Column_{i}" for i in range(1, 15)],
            )

            converter.save_raw_estimate_to_excel(df, str(output_path))

            wb = load_workbook(output_path)
            ws = wb["Estimate"]
            merged = {str(item) for item in ws.merged_cells.ranges}
            self.assertIn("E1:F1", merged)
            self.assertIn("G1:L1", merged)
            self.assertIn("I2:L2", merged)
            self.assertEqual(ws["E2"].value, "на ед.")
            self.assertEqual(ws["F2"].value, "всего")
            self.assertEqual(ws["I3"].value, "Осн.З/п")
            self.assertEqual(ws.freeze_panes, "A5")
            self.assertEqual(ws["M1"].value, "Т/з осн.\nраб.")
            self.assertEqual(ws["N1"].value, "Т/з мех.")
            self.assertEqual(ws["A5"].value, "2")
            wb.close()


if __name__ == "__main__":
    unittest.main()
