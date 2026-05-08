import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from converter import SmetaConverter
from document_export import DocumentExporter


class DocumentExportTests(unittest.TestCase):
    def test_append_header_sheet_creates_header_tab_with_image(self):
        exporter = DocumentExporter(ocr_extractor=None)

        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "book.xlsx"
            SmetaConverter().save_to_excel(pd.DataFrame({"A": ["row"]}), str(output_path))

            with patch.object(
                exporter,
                "build_header_image",
                return_value=Image.new("RGB", (320, 180), "white"),
            ):
                exporter.append_header_sheet(str(output_path), "dummy.pdf")

            wb = load_workbook(output_path)
            self.assertIn("Header", wb.sheetnames)
            ws = wb["Header"]
            self.assertFalse(ws.sheet_view.showGridLines)
            self.assertEqual(len(ws._images), 1)
            wb.close()

    def test_detect_header_bottom_uses_first_table_position_when_available(self):
        exporter = DocumentExporter(ocr_extractor=None)

        fake_page = type(
            "FakePage",
            (),
            {
                "height": 1000,
                "find_tables": lambda self: [type("T", (), {"bbox": (0, 520, 100, 900)})()],
            },
        )()

        bottom = exporter._detect_header_bottom(fake_page, 2000)

        self.assertEqual(bottom, 1010)

    def test_detect_header_bottom_falls_back_to_default_ratio_without_tables(self):
        exporter = DocumentExporter(ocr_extractor=None)

        fake_page = type(
            "FakePage",
            (),
            {
                "height": 1000,
                "find_tables": lambda self: [],
            },
        )()

        bottom = exporter._detect_header_bottom(fake_page, 2000)

        self.assertEqual(bottom, 920)


if __name__ == "__main__":
    unittest.main()
