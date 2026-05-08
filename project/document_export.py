from io import BytesIO
from typing import Optional

import cv2
import numpy as np
import pdfplumber
import pypdfium2 as pdfium
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from config import HEADER_SHEET_NAME


class DocumentExporter:
    def __init__(self, ocr_extractor, header_render_dpi: int = 180):
        self.ocr_extractor = ocr_extractor
        self.header_render_dpi = header_render_dpi

    def append_header_sheet(self, excel_path: str, pdf_path: str) -> None:
        header_image = self.build_header_image(pdf_path)
        if header_image is None:
            return

        workbook = load_workbook(excel_path)
        if HEADER_SHEET_NAME in workbook.sheetnames:
            del workbook[HEADER_SHEET_NAME]

        worksheet = workbook.create_sheet(HEADER_SHEET_NAME, 0)
        worksheet.sheet_view.showGridLines = False

        image_buffer = BytesIO()
        image_buffer.name = "header.png"
        header_image.save(image_buffer, format="PNG")
        image_buffer.seek(0)

        worksheet.add_image(XLImage(image_buffer), "A1")
        workbook.save(excel_path)
        workbook.close()

    def build_header_image(self, pdf_path: str):
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None

            first_page = pdf.pages[0]
            render_scale = self.header_render_dpi / 72
            pdf_doc = pdfium.PdfDocument(pdf_path)
            try:
                image = pdf_doc[0].render(scale=render_scale).to_pil()
            finally:
                pdf_doc.close()

            bottom = self._detect_header_bottom(first_page, image.height, image)
            bottom = max(200, min(bottom, image.height))
            return image.crop((0, 0, image.width, bottom))

    def _detect_header_bottom(self, first_page, image_height: int, rendered_image=None) -> int:
        default_bottom = int(image_height * 0.85)

        try:
            tables = first_page.find_tables()
        except Exception:
            tables = []

        if tables:
            first_table_top = min(table.bbox[1] for table in tables)
            page_height = max(float(first_page.height), 1.0)
            top_ratio = first_table_top / page_height
            margin_ratio = 0.01
            bottom = int(image_height * max(top_ratio - margin_ratio, 0.05))
            return max(bottom, 100)

        words = first_page.extract_words() or []
        if words:
            last_word_bottom = max(float(w['bottom']) for w in words)
            page_height = max(float(first_page.height), 1.0)
            bottom_ratio = last_word_bottom / page_height
            return int(image_height * min(bottom_ratio + 0.02, 1.0))

        detected_bottom = self._detect_header_bottom_from_lines(rendered_image)
        return detected_bottom or default_bottom

    def _detect_header_bottom_from_lines(self, rendered_image) -> Optional[int]:
        if rendered_image is None:
            return None

        gray = np.array(rendered_image.convert("L"))
        _, threshold = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

        v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
        vertical = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, v_kernel)
        row_has_vline = vertical.sum(axis=1) / 255
        topmost_vline = next(
            (y for y in range(len(row_has_vline)) if row_has_vline[y] >= 3),
            None,
        )
        if topmost_vline is not None:
            return max(topmost_vline - 5, 100)

        h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (80, 1))
        horizontal = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, h_kernel)
        row_sums = horizontal.sum(axis=1) / 255
        width = gray.shape[1]
        min_line_width = width * 0.30
        min_y = int(gray.shape[0] * 0.30)
        candidates = [
            row_index
            for row_index, value in enumerate(row_sums)
            if row_index >= min_y and value >= min_line_width
        ]
        if not candidates:
            return None

        first_line = candidates[0]
        return max(first_line - 5, 100)
