"""
Quality reporting helpers for PDF-to-Excel conversion.
"""

import json
import os
import re
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
import pdfplumber
from openpyxl import load_workbook

from config import (
    SOURCE_TEXT_CONTENT_LABEL,
    SOURCE_TEXT_SHEET_NAME,
    SOURCE_TEXT_SOURCE_LABEL,
)


_WHITESPACE_RE = re.compile(r"\s+", re.UNICODE)
_ILLEGAL_XML_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufffe\uffff]"
)


def _sanitize_for_excel(text: str) -> str:
    """Remove characters that openpyxl / XML cannot represent."""
    return _ILLEGAL_XML_RE.sub("", text) if text else text


def normalize_visible_text(text: str) -> str:
    return _WHITESPACE_RE.sub("", text or "").lower()


def build_text_metrics(source_text: str, excel_text: str) -> Dict[str, object]:
    source_visible = normalize_visible_text(source_text)
    excel_visible = normalize_visible_text(excel_text)

    source_counter = Counter(source_visible)
    excel_counter = Counter(excel_visible)
    matched = sum((source_counter & excel_counter).values())
    source_count = len(source_visible)
    excel_count = len(excel_visible)
    lost = max(source_count - matched, 0)

    preservation_percent = None
    if source_count:
        preservation_percent = round((matched / source_count) * 100, 2)

    return {
        "source_chars": len(source_text or ""),
        "excel_chars": len(excel_text or ""),
        "source_visible_chars": source_count,
        "excel_visible_chars": excel_count,
        "matched_visible_chars": matched,
        "lost_visible_chars": lost,
        "preservation_percent": preservation_percent,
    }


def extract_excel_text(excel_path: str) -> str:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    chunks: List[str] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    chunks.append(str(value))

    workbook.close()
    return "\n".join(chunks)


def append_source_text_sheet(excel_path: str, source_text: str, source_method: str) -> None:
    if not source_text:
        return

    workbook = load_workbook(excel_path)
    if SOURCE_TEXT_SHEET_NAME in workbook.sheetnames:
        del workbook[SOURCE_TEXT_SHEET_NAME]

    worksheet = workbook.create_sheet(SOURCE_TEXT_SHEET_NAME)
    worksheet["A1"] = SOURCE_TEXT_SOURCE_LABEL
    worksheet["B1"] = source_method
    worksheet["A2"] = SOURCE_TEXT_CONTENT_LABEL
    bold_font = copy(worksheet["A1"].font)
    bold_font.bold = True
    worksheet["A2"].font = bold_font

    safe_text = _sanitize_for_excel(source_text)
    max_cell_chars = 32767
    chunks = [
        safe_text[index:index + max_cell_chars]
        for index in range(0, len(safe_text), max_cell_chars)
    ]
    for row_index, chunk in enumerate(chunks, start=3):
        worksheet.cell(row=row_index, column=1, value=chunk)

    worksheet.column_dimensions["A"].width = 120
    workbook.save(excel_path)
    workbook.close()


def remove_source_text_sheet(excel_path: str) -> None:
    workbook = load_workbook(excel_path)
    if SOURCE_TEXT_SHEET_NAME in workbook.sheetnames:
        del workbook[SOURCE_TEXT_SHEET_NAME]
        workbook.save(excel_path)
    workbook.close()


def extract_pdf_text(pdf_path: str, use_ocr_if_needed: bool = True) -> Tuple[str, str, int]:
    text_chunks: List[str] = []
    used_ocr = False

    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        ocr_pages = None

        for page_index, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if normalize_visible_text(text):
                text_chunks.append(text)
                continue

            if not use_ocr_if_needed:
                continue

            if ocr_pages is None:
                ocr_pages = _extract_pdf_text_pages_with_ocr(pdf_path)

            if page_index < len(ocr_pages) and normalize_visible_text(ocr_pages[page_index]):
                text_chunks.append(ocr_pages[page_index])
                used_ocr = True

    combined = "\n".join(text_chunks)
    if normalize_visible_text(combined):
        return combined, "ocr_estimate" if used_ocr else "pdf_text_layer", page_count

    return "", "no_text_found", page_count


def build_quality_record(pdf_path: str, excel_path: str) -> Dict[str, object]:
    pdf_text, source_method, page_count = extract_pdf_text(pdf_path)
    remove_source_text_sheet(excel_path)
    table_text = extract_excel_text(excel_path)
    table_metrics = build_text_metrics(pdf_text, table_text)
    append_source_text_sheet(excel_path, pdf_text, source_method)
    excel_text = extract_excel_text(excel_path)
    metrics = build_text_metrics(pdf_text, excel_text)

    warning = ""
    if source_method == "ocr_estimate":
        warning = "PDF has no text layer; the source count is based on OCR."
    elif source_method == "no_text_found":
        warning = "Could not read any source text from the PDF."

    return {
        "pdf_file": Path(pdf_path).name,
        "excel_file": Path(excel_path).name,
        "pages": page_count,
        "source_method": source_method,
        "table_visible_chars": table_metrics["excel_visible_chars"],
        "table_matched_visible_chars": table_metrics["matched_visible_chars"],
        "table_lost_visible_chars": table_metrics["lost_visible_chars"],
        "table_preservation_percent": table_metrics["preservation_percent"],
        **metrics,
        "warning": warning,
    }


def save_quality_reports(records: Iterable[Dict[str, object]], output_dir: str) -> Dict[str, str]:
    rows = list(records)
    os.makedirs(output_dir, exist_ok=True)

    json_path = str(Path(output_dir) / "conversion_report.json")
    xlsx_path = str(Path(output_dir) / "conversion_report.xlsx")

    with open(json_path, "w", encoding="utf-8") as report_file:
        json.dump(rows, report_file, ensure_ascii=False, indent=2)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame([{"message": "No converted files to report."}])
    df.to_excel(xlsx_path, index=False)

    return {"json": json_path, "xlsx": xlsx_path}


def build_reports_for_pairs(pairs: Iterable[Tuple[str, str]], output_dir: str) -> Dict[str, object]:
    records = []
    for pdf_path, excel_path in pairs:
        try:
            records.append(build_quality_record(pdf_path, excel_path))
        except Exception as exc:
            records.append(
                {
                    "pdf_file": Path(pdf_path).name,
                    "excel_file": Path(excel_path).name,
                    "error": str(exc),
                }
            )

    paths = save_quality_reports(records, output_dir)
    return {"records": records, "paths": paths}


def summarize_quality(records: Iterable[Dict[str, object]]) -> str:
    rows = [row for row in records if "preservation_percent" in row]
    if not rows:
        return "Quality report: no comparable rows were generated."

    comparable = [row for row in rows if row.get("preservation_percent") is not None]
    if not comparable:
        return "Quality report: the PDFs do not contain a readable text layer, so no exact percentage was computed."

    average = round(
        sum(float(row["preservation_percent"]) for row in comparable) / len(comparable),
        2,
    )
    worst = min(comparable, key=lambda row: float(row["preservation_percent"]))
    return (
        f"Quality report: average preservation {average}%, "
        f"worst file {worst['pdf_file']} at {worst['preservation_percent']}% "
        f"({worst['lost_visible_chars']} visible characters missing in Excel)."
    )


def _extract_pdf_text_pages_with_ocr(pdf_path: str) -> List[str]:
    import shutil

    try:
        import pypdfium2 as pdfium
        import pytesseract
    except Exception:
        return []

    _path = shutil.which("tesseract")
    if _path:
        pytesseract.pytesseract.tesseract_cmd = _path
    elif os.path.exists(r"C:\Program Files\Tesseract-OCR\tesseract.exe"):
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

    chunks: List[str] = []
    try:
        pdf_doc = pdfium.PdfDocument(pdf_path)
        for page in pdf_doc:
            bitmap = page.render(scale=300 / 72)
            image = bitmap.to_pil()
            text = pytesseract.image_to_string(image, lang="rus")
            chunks.append(text)
        pdf_doc.close()
    except Exception:
        return []

    return chunks
