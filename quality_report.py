"""
Quality report for PDF -> Excel conversion.

The report answers the practical client question: how many visible
characters from the PDF text layer are present in the resulting XLSX.
For scanned PDFs without a text layer the source text can only be
estimated with OCR, so the report marks that separately.
"""

import json
import os
import re
from copy import copy
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import pdfplumber
from openpyxl import load_workbook


_WHITESPACE_RE = re.compile(r"\s+", re.UNICODE)


def normalize_visible_text(text: str) -> str:
    """Return comparable visible text: lowercase, without whitespace."""
    return _WHITESPACE_RE.sub("", text or "").lower()


def build_text_metrics(source_text: str, excel_text: str) -> Dict[str, object]:
    """Build preservation metrics using visible characters, ignoring order."""
    source_visible = normalize_visible_text(source_text)
    excel_visible = normalize_visible_text(excel_text)

    source_counter = Counter(source_visible)
    excel_counter = Counter(excel_visible)
    matched = sum((source_counter & excel_counter).values())
    source_count = len(source_visible)
    excel_count = len(excel_visible)
    lost = max(source_count - matched, 0)

    preservation_percent = None
    if source_count:
        preservation_percent = round((matched / source_count) * 100, 2)

    return {
        "source_chars": len(source_text or ""),
        "excel_chars": len(excel_text or ""),
        "source_visible_chars": source_count,
        "excel_visible_chars": excel_count,
        "matched_visible_chars": matched,
        "lost_visible_chars": lost,
        "preservation_percent": preservation_percent,
    }


def extract_excel_text(excel_path: str) -> str:
    """Read all cell values from all sheets in an XLSX file as plain text."""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    chunks: List[str] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    chunks.append(str(value))

    workbook.close()
    return "\n".join(chunks)


def append_source_text_sheet(excel_path: str, source_text: str, source_method: str) -> None:
    """Add or replace a sheet with full source text for auditability."""
    if not source_text:
        return

    workbook = load_workbook(excel_path)
    sheet_name = "Исходный текст"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(sheet_name)
    worksheet["A1"] = "Источник"
    worksheet["B1"] = source_method
    worksheet["A2"] = "Текст PDF"
    bold_font = copy(worksheet["A1"].font)
    bold_font.bold = True
    worksheet["A2"].font = bold_font

    max_cell_chars = 32767
    chunks = [
        source_text[index:index + max_cell_chars]
        for index in range(0, len(source_text), max_cell_chars)
    ]
    for row_index, chunk in enumerate(chunks, start=3):
        worksheet.cell(row=row_index, column=1, value=chunk)

    worksheet.column_dimensions["A"].width = 120
    workbook.save(excel_path)
    workbook.close()


def remove_source_text_sheet(excel_path: str) -> None:
    """Remove the audit source sheet when recalculating table-only metrics."""
    workbook = load_workbook(excel_path)
    sheet_name = "Исходный текст"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
        workbook.save(excel_path)
    workbook.close()


def extract_pdf_text(pdf_path: str, use_ocr_if_needed: bool = True) -> Tuple[str, str, int]:
    """Extract comparable source text from a PDF and return text, method, pages."""
    text_chunks: List[str] = []
    page_count = 0

    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                text_chunks.append(text)

    text_layer = "\n".join(text_chunks)
    if normalize_visible_text(text_layer) or not use_ocr_if_needed:
        return text_layer, "pdf_text_layer", page_count

    ocr_text = _extract_pdf_text_with_ocr(pdf_path)
    if normalize_visible_text(ocr_text):
        return ocr_text, "ocr_estimate", page_count

    return "", "no_text_found", page_count


def build_quality_record(pdf_path: str, excel_path: str) -> Dict[str, object]:
    """Build one quality report row for a converted PDF/XLSX pair."""
    pdf_text, source_method, page_count = extract_pdf_text(pdf_path)
    remove_source_text_sheet(excel_path)
    table_text = extract_excel_text(excel_path)
    table_metrics = build_text_metrics(pdf_text, table_text)
    append_source_text_sheet(excel_path, pdf_text, source_method)
    excel_text = extract_excel_text(excel_path)
    metrics = build_text_metrics(pdf_text, excel_text)

    warning = ""
    if source_method == "ocr_estimate":
        warning = "PDF has no text layer; source count is OCR estimate."
    elif source_method == "no_text_found":
        warning = "Could not read source text from PDF."

    return {
        "pdf_file": Path(pdf_path).name,
        "excel_file": Path(excel_path).name,
        "pages": page_count,
        "source_method": source_method,
        "table_visible_chars": table_metrics["excel_visible_chars"],
        "table_matched_visible_chars": table_metrics["matched_visible_chars"],
        "table_lost_visible_chars": table_metrics["lost_visible_chars"],
        "table_preservation_percent": table_metrics["preservation_percent"],
        **metrics,
        "warning": warning,
    }


def save_quality_reports(records: Iterable[Dict[str, object]], output_dir: str) -> Dict[str, str]:
    """Save quality records to JSON and XLSX files."""
    rows = list(records)
    os.makedirs(output_dir, exist_ok=True)

    json_path = str(Path(output_dir) / "conversion_report.json")
    xlsx_path = str(Path(output_dir) / "conversion_report.xlsx")

    with open(json_path, "w", encoding="utf-8") as report_file:
        json.dump(rows, report_file, ensure_ascii=False, indent=2)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame([{"message": "No converted files to report."}])
    df.to_excel(xlsx_path, index=False)

    return {"json": json_path, "xlsx": xlsx_path}


def build_reports_for_pairs(pairs: Iterable[Tuple[str, str]], output_dir: str) -> Dict[str, object]:
    """Build and save a report for many PDF/XLSX output pairs."""
    records = []
    for pdf_path, excel_path in pairs:
        try:
            records.append(build_quality_record(pdf_path, excel_path))
        except Exception as exc:
            records.append({
                "pdf_file": Path(pdf_path).name,
                "excel_file": Path(excel_path).name,
                "error": str(exc),
            })

    paths = save_quality_reports(records, output_dir)
    return {"records": records, "paths": paths}


def summarize_quality(records: Iterable[Dict[str, object]]) -> str:
    """Return a short human-readable report summary."""
    rows = [row for row in records if "preservation_percent" in row]
    if not rows:
        return "Отчет качества: нет данных для сравнения."

    comparable = [row for row in rows if row.get("preservation_percent") is not None]
    if not comparable:
        return "Отчет качества: PDF без читаемого текстового слоя, точный процент не посчитан."

    avg = round(
        sum(float(row["preservation_percent"]) for row in comparable) / len(comparable),
        2,
    )
    worst = min(comparable, key=lambda row: float(row["preservation_percent"]))
    return (
        f"Отчет качества: средняя сохранность {avg}%, "
        f"хуже всего {worst['pdf_file']} - {worst['preservation_percent']}% "
        f"({worst['lost_visible_chars']} видимых символов не найдено в Excel)."
    )


def _extract_pdf_text_with_ocr(pdf_path: str) -> str:
    """Best-effort OCR fallback for scanned PDFs."""
    try:
        import pypdfium2 as pdfium
        import pytesseract
    except Exception:
        return ""

    tesseract_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

    chunks: List[str] = []
    try:
        pdf_doc = pdfium.PdfDocument(pdf_path)
        for page in pdf_doc:
            bitmap = page.render(scale=300 / 72)
            image = bitmap.to_pil()
            chunks.append(pytesseract.image_to_string(image, lang="rus+eng"))
    except Exception:
        return ""

    return "\n".join(chunks)
